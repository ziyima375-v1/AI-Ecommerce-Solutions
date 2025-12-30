# -*- coding: utf-8 -*-
"""
电商素材表 一键生成（整合版 · 失败剔除日志+自动重置索引+双栏进度UI+批量队列）
【本版优化要点】
1) 删除 DB/BN/BT 的任何写入，完全以模板为准（DB 不再写主图；BN 全部清空；BT 不写）。
2) EQ 固定“欧盟尺码常规”；FC/FD/FE=196/192/185。
3) 轮播图：每个站点第1张写入上传URL，其余4张用模板既定4张；
   且仅从“第2件”开始执行，第一件（1个SPU+6个SKU）完全保留模板的轮播列。
   同时从第2件起，SPU 的 GV 以及 IC/IN/IY/JJ/JU/KF/KQ 写成全角“｜轮播1｜轮播2｜…｜”。
4) 若 SPU 行的 ET/FF 经过复制后出现空白，自动回填为模板该列值。
除以上显式改动点外，其余功能代码 100% 保留。
"""
import os
import re
import json
import time
import base64
import eel
import random
import pandas as pd
from io import BytesIO
from PIL import Image
from datetime import datetime
from threading import Semaphore, Thread
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from copy import copy

# --------------------------- 路径与默认配置 ---------------------------
APP_DIR = os.path.dirname(os.path.abspath(__file__))

# 前端资源放在程序目录（不要放桌面）
WEB_DIR = os.path.join(APP_DIR, "web")
os.makedirs(WEB_DIR, exist_ok=True)

# 唯一桌面目录与产出子目录
APP_NAME = "Ecom-Excel-Tool"
SUB_OUT = "输出"

def _desktop_data_dir():
    base = os.path.join(os.path.expanduser("~"), "Desktop", APP_NAME)
    os.makedirs(base, exist_ok=True)
    return base

DATA_DIR = _desktop_data_dir()

# 所有持久化与产出都放到 DATA_DIR 下
CONFIG_TITLE = os.path.join(DATA_DIR, "title_config.json")
FORBIDDEN_FILE = os.path.join(DATA_DIR, "forbidden_words.txt")
EXCEL_CFG = os.path.join(DATA_DIR, "excel_tool_config.json")
OUT_DIR = os.path.join(DATA_DIR, SUB_OUT)
os.makedirs(OUT_DIR, exist_ok=True)

DEFAULT_PROMPT = (
    "You are a precise title writer for US e-commerce. "
    "Write ONE single-line product title only (no extra text). "
    "Style must fit mainstream US aesthetics: eye-catching, concise, unique. "
    "Hard rules:\n"
    "- Include design style, key visual elements, and theme.\n"
    "- Include 2-3 high-demand US keywords (e.g., comfortable cotton, casual wear, party essential, graphic tee).\n"
    "- Mention '100% cotton' and comfort in a 15-20 word sentence.\n"
    "- Keep the body UNDER 230 characters (reserve space for an ending phrase).\n"
    "- End with exactly one of: Perfect for gifting OR Ideal for casual wear.\n"
    "- No quotes or apostrophes. Output a single line only."
)
# ==== 样式控制 ====
COPY_TEMPLATE_STYLE = False          # False=不复制模板样式（只复制值）
CLEAR_STYLES_AFTER_INSERT = True     # True=插入完成后，整块清空边框/底纹/对齐等

# ==== 行为开关 ====
# True: 第1件（1个SPU+6个SKU）的所有轮播列保留模板原样；
# 从第2件起才把SKU轮播第1张写入上传URL、2~5张写模板固定4张，并写 SPU 的合并列。
KEEP_FIRST_ITEM_CAROUSEL_AS_TEMPLATE = False

eel.init(WEB_DIR)

# --------------------------- 模块一：图生标题 ---------------------------
class WindowTitleGenerator:
    def __init__(self):
        self.api_key = ""  # 对外发布请通过环境变量/本地文件配置密钥（见下方注释）
        env_key = (os.getenv("DASHSCOPE_API_KEY") or os.getenv("QWEN_API_KEY") or "").strip()
        if env_key:
            self.api_key = env_key
        else:
            key_file = os.path.join(DATA_DIR, "api_key.txt")
            if os.path.exists(key_file):
                try:
                    self.api_key = open(key_file, "r", encoding="utf-8").read().strip()
                except Exception:
                    self.api_key = ""
        self.base_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"

        self.max_retries = 3
        self.initial_timeout = 10
        self.backoff_factor = 0.5

        self.session = requests.Session()
        adapter = HTTPAdapter(
            pool_connections=20, pool_maxsize=20,
            max_retries=Retry(total=2, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
        self.common_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Cache-Control": "no-cache"
        }

        self.max_workers = 8
        self.api_sema = Semaphore(5)

        self.image_urls = pd.DataFrame()
        self.url_column = ""
        self.results = []
        self.generating = False

        self.prompt_content = self._load_prompt()
        self.forbidden_words = self._load_forbidden()
        self.save_file_path = ""
        self.current_excel_path = ""

    # 提示词/违禁词
    def _load_prompt(self):
        if os.path.exists(CONFIG_TITLE):
            try:
                cfg = json.load(open(CONFIG_TITLE, "r", encoding="utf-8"))
                return cfg.get("prompt", DEFAULT_PROMPT)
            except Exception:
                return DEFAULT_PROMPT
        return DEFAULT_PROMPT

    def save_prompt(self, content):
        content = (content or "").strip() or DEFAULT_PROMPT
        data = {"prompt": content}
        try:
            if os.path.exists(CONFIG_TITLE):
                old = json.load(open(CONFIG_TITLE, "r", encoding="utf-8"))
                old.update(data)
                data = old
        except Exception:
            pass
        with open(CONFIG_TITLE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        self.prompt_content = content
        eel.show_message("提示词", "提示词已保存")()
        eel.update_prompt_preview(self.prompt_content)()

    def _load_forbidden(self):
        if os.path.exists(FORBIDDEN_FILE):
            try:
                return self.parse_forbidden_words(open(FORBIDDEN_FILE, "r", encoding="utf-8").read())
            except Exception:
                return []
        return []

    def save_forbidden_words(self, content):
        self.forbidden_words = self.parse_forbidden_words(content or "")
        with open(FORBIDDEN_FILE, "w", encoding="utf-8") as f:
            f.write(", ".join(self.forbidden_words))
        eel.show_message("违禁词", f"违禁词已保存，共 {len(self.forbidden_words)} 个")()
        eel.update_forbidden_words(self.forbidden_words)()

    # 载入 素材采集 Excel
    def load_temu_excel(self, file_path: str):
        if not file_path or not os.path.exists(file_path):
            eel.show_warning("文件错误", "请选择正确的 素材采集 素材 Excel")()
            return {"status": "error", "message": "未选择有效Excel"}
        try:
            df = pd.read_excel(file_path)
            url_cols = [c for c in df.columns if
                        "url" in str(c).lower() or "图片" in str(c) or "image" in str(c).lower()]
            if not url_cols:
                eel.show_warning("未找到URL", "Excel中没有发现包含 URL 的列（url/图片/image）")()
                return {"status": "error", "message": "未找到URL列"}
            self.url_column = url_cols[0]
            self.image_urls = df[[self.url_column]].dropna().reset_index()
            self.image_urls.columns = ["original_index", self.url_column]
            self.current_excel_path = file_path
            eel.show_message("Excel", f"已读取 {len(self.image_urls)} 个图片URL（列：{self.url_column}）")()
            eel.update_status(f"URL 待处理：{len(self.image_urls)}")()
            try:
                cfg = {}
                if os.path.exists(CONFIG_TITLE):
                    cfg = json.load(open(CONFIG_TITLE, "r", encoding="utf-8"))
                cfg["last_temu_excel"] = file_path
                json.dump(cfg, open(CONFIG_TITLE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            except Exception:
                pass
            return {"status": "success", "count": len(self.image_urls)}
        except Exception as e:
            eel.show_error(f"读取Excel失败: {e}")()
            return {"status": "error", "message": str(e)}

    # 主流程：并发生成标题
    def process_all(self):
        if self.image_urls.empty:
            eel.show_warning("缺少数据", "请先选择 素材采集 Excel")()
            return {"status": "error", "message": "无URL"}
        if not self.prompt_content:
            eel.show_warning("缺少提示词", "请先设置提示词")()
            return {"status": "error", "message": "无提示词"}

        self.results = []
        self.generating = True
        eel.update_progress(0)()
        total = len(self.image_urls)

        tasks = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
            for i in range(total):
                row = self.image_urls.iloc[i]
                tasks.append(ex.submit(self._process_one, i, row))
            done = 0
            for fut in as_completed(tasks):
                try:
                    result = fut.result()
                except Exception as e:
                    result = {"original_index": -1, "image_url": "", "title": f"生成失败: {e}", "has_forbidden": False,
                              "matched_words": []}
                self.results.append(result)
                done += 1
                if done % 5 == 0 or done == total:
                    eel.update_progress(done * 100 / total)()
                    eel.update_status(f"标题生成 {done}/{total}")()

        self.generating = False
        bridge_path, ok_count = self._save_bridge_excel()
        eel.update_status("标题已生成，准备生成上架表格")()
        return {"status": "success", "bridge": bridge_path, "count": ok_count}

    def _process_one(self, idx, row):
        url = row[self.url_column]
        b64, _small = self._download_and_shrink(url)
        with self.api_sema:
            raw = self._call_llm_api(b64, url)
        title = self.extract_title(raw)
        matched = self.contains_forbidden_word(title)
        return {"original_index": row["original_index"], "image_url": url, "title": title,
                "has_forbidden": bool(matched), "matched_words": matched}

    def _download_and_shrink(self, url, max_side=1024, quality=85):
        last_exc = None
        for attempt in range(self.max_retries):
            try:
                resp = self.session.get(url, timeout=min(self.initial_timeout, 10), headers=self.common_headers,
                                        stream=True, verify=True)
                resp.raise_for_status()
                raw = resp.content
                with Image.open(BytesIO(raw)) as im:
                    im = im.convert("RGB")
                    im.thumbnail((max_side, max_side))
                    buf = BytesIO()
                    im.save(buf, format="JPEG", quality=85, optimize=True)
                    small = buf.getvalue()
                return base64.b64encode(small).decode("utf-8"), small
            except Exception as e:
                last_exc = e
                time.sleep(self.backoff_factor * (attempt + 1))
        raise Exception(f"下载图片失败: {last_exc}")

    def _call_llm_api(self, base64_image, image_url):
        fmt = self._get_format_from_url(image_url)
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {self.api_key}"}
        data = {
            "model": "qwen-vl-max",
            "messages": [
                {"role": "system", "content": DEFAULT_PROMPT},
                {"role": "user", "content": [
                    {"type": "text", "text": self.prompt_content},
                    {"type": "image_url", "image_url": {"url": f"data:image/{fmt};base64,{base64_image}"}}
                ]}
            ],
            "temperature": 0.5, "max_tokens": 140
        }
        r = self.session.post(f"{self.base_url}/chat/completions", headers=headers, json=data, timeout=25)
        r.raise_for_status()
        j = r.json()
        if "choices" in j and j["choices"]:
            return j["choices"][0]["message"]["content"].strip()
        return "未能生成标题，请重试"

    def _get_format_from_url(self, url):
        l = (url or "").lower()
        for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
            if l.endswith(ext): return ext[1:]
        return "jpeg"

    # 清洗 & 违禁词
    def extract_title(self, raw_text):
        if not raw_text: return "No valid title generated"
        import re as _re
        english_pattern = _re.compile(r'[a-zA-Z0-9\s.,!?()\-]+')
        parts = english_pattern.findall(raw_text)
        if not parts: return "No English content detected"
        title = _re.sub(r'\s+', ' ', ' '.join(parts)).strip()
        title = title.replace('"', '').replace("'", "")
        BODY_MAX, MAX_LEN = 230, 250
        end_opts = ["Perfect for gifting", "Ideal for casual wear"]
        chosen = min(end_opts, key=len)
        for e in end_opts:
            if e.lower() in title.lower():
                chosen = e
                title = _re.sub(_re.escape(e), '', title, flags=_re.IGNORECASE).strip()
        if len(title) > BODY_MAX:
            cut = title[:BODY_MAX]
            ls = cut.rfind(' ')
            title = (cut[:ls] if ls > 0 else cut).rstrip(' ,.!?()-')
        reserve = 1 + len(chosen)
        if len(title) + reserve > MAX_LEN:
            new_body_max = MAX_LEN - reserve
            if new_body_max < 0: return chosen
            if len(title) > new_body_max:
                cut = title[:new_body_max]
                ls = cut.rfind(' ')
                title = (cut[:ls] if ls > 0 else cut).rstrip(' ,!?()-')
        final = f"{title} {chosen}" if title else chosen
        return final[:MAX_LEN].rstrip()

    def parse_forbidden_words(self, content):
        if not content: return []
        parts = [x.strip().lower() for x in content.split(',') if x.strip()]
        seen, out = set(), []
        for w in parts:
            if w not in seen:
                seen.add(w); out.append(w)
        return out

    def contains_forbidden_word(self, title):
        if not title or not self.forbidden_words: return []
        tl = title.lower()
        matched = []
        for w in self.forbidden_words:
            if not w or len(w) < 3: continue
            pattern = r'(^|\s|,|\.|;|!|\?|/|\\|\(|\))' + re.escape(w) + r'($|\s|,|\.|;|!|\?|/|\\|\(|\))'
            if re.search(pattern, tl): matched.append(w)
        if matched: return list(set(matched))
        brand_keywords = ['nike', 'adidas', 'jordan', 'puma', 'gucci', 'lv', 'chanel', 'dior', 'hermes', 'burberry',
                          'balenciaga', 'versace', 'fendi', 'prada']
        for w in [w for w in self.forbidden_words if w in brand_keywords]:
            if len(w) > 4 and w[:4] in tl and re.search(r'\b' + re.escape(w[:4]) + r'\w*', tl):
                matched.append(w)
        return list(set(matched))

    def _save_bridge_excel(self):
        if not self.results:
            return "", 0
        def ok(r):
            t = (r.get("title") or "").strip()
            return (t and not t.startswith("生成失败") and t != "未能生成标题，请重试" and r.get("image_url"))
        ok_results = [r for r in self.results if ok(r)]
        sorted_results = sorted(ok_results, key=lambda r: r.get("original_index", 0))
        df = pd.DataFrame({
            "生成的标题": [r["title"] for r in sorted_results],
            "图片URL": [r["image_url"] for r in sorted_results],
            "是否含违禁词": ["是" if r.get("has_forbidden", False) else "否" for r in sorted_results],
            "违禁词": [", ".join(r.get("matched_words", [])) for r in sorted_results],
        })
        fn = os.path.join(OUT_DIR, f"标题中转_{datetime.now().strftime('%m%d%H%M')}.xlsx")
        df.to_excel(fn, index=False)
        self.save_file_path = fn
        eel.update_save_path(fn)()
        return fn, len(sorted_results)

    # === 英文→多语种翻译 ===
    
    # === 英文→多语种翻译（更稳健：多路兜底） ===
    def translate_multi(self, english_title: str) -> dict:
        english_title = (english_title or "").strip()
        if not english_title:
            return {}
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {self.api_key}"}
        sys = (
            "You are a professional e-commerce localization translator. "
            "Translate the provided English product title into the requested languages. "
            "Output strict JSON with keys: de, fr, it, es, cs, uk, pl, pt, hu, nl, sv. "
            "Keep it concise (<= 150 chars) and natural for product titles; no extra commentary."
        )
        user_text = (
            "Title:\\n" + english_title +
            "\\nLanguages: German(de), French(fr), Italian(it), Spanish(es), Czech(cs), Ukrainian(uk), "
            "Polish(pl), Portuguese(pt), Hungarian(hu), Dutch(nl), Swedish(sv). "
            "Return pure JSON only."
        )
        data = {
            "model": "qwen-turbo",
            "messages": [
                {"role": "system", "content": sys},
                {"role": "user", "content": user_text}
            ],
            "temperature": 0.2,
            "max_tokens": 400
        }
        lang_keys = ['de','fr','it','es','cs','uk','pl','pt','hu','nl','sv']
        result = {k: english_title for k in lang_keys}
        try:
            r = self.session.post(f"{self.base_url}/chat/completions", headers=headers, json=data, timeout=25)
            r.raise_for_status()
            j = r.json()
            content = j.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            import json as _json, re as _re
            m = _re.search(r'\\{.*\\}', content, _re.S)
            js = _json.loads(m.group(0) if m else content)
            # 先用批量结果
            for k in lang_keys:
                v = (js.get(k,"") or "").strip()
                if v:
                    result[k] = v
        except Exception:
            pass

        # 逐语种补救：如果仍是英文或者与英文高度一致，则单独翻译一次
        for k in lang_keys:
            v = (result.get(k, "") or "").strip()
            if not v or v.lower() == english_title.lower():
                try:
                    result[k] = self._translate_one(english_title, k) or english_title
                except Exception:
                    result[k] = english_title
        return result

    def _translate_one(self, english_title: str, lang_key: str) -> str:
        # 单语种兜底翻译：只返回纯文本
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {self.api_key}"}
        name_map = {
            'de': 'German', 'fr': 'French', 'it': 'Italian', 'es': 'Spanish',
            'cs': 'Czech', 'uk': 'Ukrainian', 'pl': 'Polish', 'pt': 'Portuguese',
            'hu': 'Hungarian', 'nl': 'Dutch', 'sv': 'Swedish'
        }
        sys = (
            "You are a professional e-commerce localization translator. "
            f"Translate the user provided English product title into {name_map.get(lang_key, lang_key)}. "
            "Return ONLY the translated title text, no quotes, no extra characters."
        )
        data = {
            "model": "qwen-turbo",
            "messages": [
                {"role": "system", "content": sys},
                {"role": "user", "content": english_title}
            ],
            "temperature": 0.2,
            "max_tokens": 150
        }
        r = self.session.post(f"{self.base_url}/chat/completions", headers=headers, json=data, timeout=20)
        r.raise_for_status()
        j = r.json()
        text = (j.get("choices", [{}])[0].get("message", {}).get("content", "") or "").strip()
        # 去除可能的包裹代码块/引号
        text = re.sub(r'^["\\s`]+|["\\s`]+$', '', text)
        return text
# --------------------------- 模块二：生成上架表格 ---------------------------
class ExcelToolApp:
    def __init__(self):
        self.config_file = EXCEL_CFG
        self.spu_history_file = os.path.join(DATA_DIR, "spu_history.txt")
        self._init_config()

        self.file_path = ""
        self.title_image_file_path = ""
        self.selected_sheet = "模版"
        self.selected_title_image_sheet = ""
        self.sheets = []
        self.title_image_sheets = []

        self.template_start = 6
        self.template_end = 12
        self.title_row = self.template_start
        self.image_column = column_index_from_string("DB")   # ← 仍记录，但不再写入
        self.english_title_col_C = column_index_from_string("C")
        self.english_title_col_D = column_index_from_string("D")

        self.dl_column = column_index_from_string("DL")      # 历史兼容（不再强制写）
        self.bn_column = column_index_from_string("BN")      # BN 改为清空
        self.gv_column = column_index_from_string("GV")      # SPU 汇总轮播列
        # SPU 的“轮播汇总”列（全部写成｜轮播1｜…｜）
        self.agg_carousel_cols = ["HG", "HR", "IC", "IN", "IY", "JJ", "JU", "KF", "KQ"]

        self.sizes = ["S", "M", "L", "XL", "XXL", "XXXL"]

        self.titles_list = []
        self.english_titles_list = []
        self.images_list = []
        self.logs = []
        self.template_df = None
        self.template_max_col = 0

        # 语言列映射（SPU 行）
        self.lang_map = {
            'de': 'E', 'fr': 'F', 'it': 'G', 'es': 'H', 'cs': 'I',
            'uk': 'J', 'uk_dup': 'K', 'pl': 'L', 'pt': 'M', 'hu': 'N',
            'nl': 'O', 'nl_dup': 'P', 'sv': 'Q',
        }

        # SPU 行固定字段（保留）
        self.spu_fixed = {
            'R': "意大利",
            'T': "棉",
            'U': "棉",
            'V': "100%",
            'AE': "复古",
            'AF': "可机洗且可干洗",
            'AG': "上衣",
            'AH': "圆领",
            'AI': "微弹",
            'AJ': "印花",
            'AK': "贴花",
            'AL': "休闲",
            'AV': "男士",
            'BC': "春,夏,秋,冬",
            'BL': "运动",
            'BV': "常规",
            'BW': "针织(含钩织、毛织面料)",
            'BX': "定位印花",
            'BY': "光面",
            'BZ': 125,
            'CA': "g/㎡",
            'CB': "光面",
            'CC': 125,
            'CD': "g/㎡",
            'CE': "棉Cotton",
            'CF': "100%",
            'EY': "Owen",
            'EZ': "XL",
            'FA': "合身",
        }

        # SKU 行固定字段模板
        self.sku_cols = {
            'EN': "黑色",
            'EO': "尺码",
            'EP': "欧美尺码",
            'EQ': "欧盟尺码常规",  # ← 对齐要求
        }
        self.sku_size_cols = {
            'ER': ["S", "M", "L", "XL", "XXL", "XXXL"],
            'ES': [36, 38, 40, 42, 44, 46],
            # 'BT': [...],  # ← 已删除 BT 的写入
            'EU': [92, 102, 112, 122, 132, 142],
            'EV': [71, 74, 76, 79, 82, 84],
            'EW': [21, 21, 23, 23, 25, 25],
            'EX': [92, 102, 112, 122, 132, 142],
            'ET': [43, 45, 48, 51, 53, 56],
        }
        # FC/FD/FE 等固定值
        self.sku_same_for_all_rows = {
            'FC': 196, 'FD': 192, 'FE': 185,
            'FG': 206, 'FH': 197, 'FI': 197,
            'FJ': 228, 'FK': 206, 'FL': 206, 'FM': 228,
            'FN': "CNY", 'FO': "极特意大利", 'FP': 20000, 'FQ': "单品", 'FR': 1, 'FS': "件",
            'FU': 268, 'FV': "CNY",
            'FF': 192,  # ← 参照你的模板反馈
            'GH': 30, 'GI': 25, 'GJ': 1, 'GK': 220
        }

        # 各站点轮播起始列（每站 5 列）
        self.station_starts = ["GL", "GW", "HH", "HS", "ID", "IO", "IZ", "JK", "JV", "KG"]

        # 轮播图2~5 固定模板 URL
        self.carousel_fixed_2_to_5 = [
            "https://img.kwcdn.com/product/fancy/9ba3044e-308a-473c-8c0a-d095d3d076c3.jpg",  # GM
            "https://img.kwcdn.com/product/fancy/015f812a-4f93-4321-9647-1d3f25aadc04.jpg",  # GN
            "https://img.kwcdn.com/product/fancy/f2845c74-378d-4a51-9dd8-8b6fe59162f6.jpg",  # GO
            "https://img.kwcdn.com/product/fancy/25b7dd66-3527-45b4-b578-d1b97dab6d9e.jpg",  # GP
        ]

    def _init_config(self):
        default = {
            "last_title_index": 0,
            "used_titles": [],
            "last_image_index": 0,
            "used_images": [],
            "last_spu_start": "",
            "last_spu_end": "",
            "default_template_path": ""
        }
        if os.path.exists(self.config_file):
            try:
                self.config = json.load(open(self.config_file, "r", encoding="utf-8"))
                for k, v in default.items():
                    if k not in self.config: self.config[k] = v
            except Exception:
                self.config = default
        else:
            self.config = default
            json.dump(self.config, open(self.config_file, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    def _save_config(self):
        json.dump(self.config, open(self.config_file, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    def reset_indices(self):
        self.config["last_title_index"] = 0
        self.config["last_image_index"] = 0
        self._save_config()
        self._log("已自动重置索引为 0 / 0。")

    def _calc_next_spu_start(self):
        last_end = (self.config or {}).get("last_spu_end", "")
        m = re.match(r'^([A-Za-z]+)(\d+)$', last_end or "")
        if not m:
            return "CS00001"
        prefix, num = m.group(1), int(m.group(2))
        width = len(m.group(2))
        return f"{prefix}{str(num + 1).zfill(width)}"

    def _log(self, msg, is_error=False):
        self.logs.append({"type": "error" if is_error else "info", "message": msg})
        if len(self.logs) > 1000: self.logs = self.logs[-1000:]
        eel.update_logs(self.logs)()

    def _save_spu_history(self, template_filename, spu_start, spu_end):
        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            line = f"[{current_time}] 模板: {template_filename} | SPU: {spu_start} - {spu_end}\n"
            with open(self.spu_history_file, "a", encoding="utf-8") as f:
                f.write(line)
        except Exception as e:
            self._log(f"保存SPU历史失败: {e}", is_error=True)

    # 载入模板与桥接表
    def load_target_file(self, file_path):
        try:
            self.file_path = file_path
            self._log(f"加载模板文件: {os.path.basename(file_path)}")
            wb = load_workbook(file_path, read_only=True)
            if "模版" not in wb.sheetnames:
                all_sheets = wb.sheetnames; wb.close()
                return {"status": "error", "message": f"未找到'模版'工作表；文件包含：{', '.join(all_sheets)}"}
            sheet = wb["模版"]
            self.template_max_col = sheet.max_column
            wb.close()

            self.template_df = pd.read_excel(
                file_path,
                sheet_name="模版",
                header=None,
                skiprows=self.template_start - 1,
                nrows=self.template_end - self.template_start + 1,
                usecols=range(self.template_max_col)
            )

            self.sheets = ["模版"]
            self.config["default_template_path"] = file_path
            self._save_config()
            return {"status": "success", "message": "模板已加载"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def load_title_image_file(self, file_path):
        try:
            self.title_image_file_path = file_path
            excel_file = pd.ExcelFile(file_path)
            self.title_image_sheets = excel_file.sheet_names
            if not self.title_image_sheets:
                return {"status": "error", "message": "标题图片文件没有工作表"}
            self.selected_title_image_sheet = self.title_image_sheets[0]
            return self._load_titles_images_from_excel()
        except Exception as e:
            return {"status": "error", "message": str(e)}

    def _load_titles_images_from_excel(self):
        if not self.title_image_file_path:
            return {"status": "error", "message": "未选择标题图片Excel"}
        try:
            df = pd.read_excel(self.title_image_file_path, sheet_name=self.selected_title_image_sheet, header=None)
            self.titles_list, self.english_titles_list, self.images_list = [], [], []
            for i in range(1, len(df)):  # 跳过表头
                title_val = df.iloc[i, 0] if 0 < len(df.columns) else ""
                image_val = df.iloc[i, 1] if 1 < len(df.columns) else ""
                if title_val and str(title_val).strip():
                    self.titles_list.append(str(title_val).strip())
                    self.english_titles_list.append(str(title_val).strip())
                if image_val and str(image_val).strip():
                    self.images_list.append(str(image_val).strip())
            return {"status": "success", "message": "已载入标题/图片",
                    "titles_count": len(self.titles_list), "images_count": len(self.images_list),
                    "last_title_index": self.config["last_title_index"],
                    "last_image_index": self.config["last_image_index"]}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    # 模板缓存
    def _apply_template_row_cached(self, ws, target_row, row_idx, style_cache, value_cache):
        # 仅按需要复制样式
        if COPY_TEMPLATE_STYLE:
            row_style = style_cache[row_idx]
            for col, st in row_style.items():
                cell = ws.cell(row=target_row, column=col)
                if st["font"] is not None: cell.font = st["font"]
                if st["fill"] is not None: cell.fill = st["fill"]
                if st["border"] is not None: cell.border = st["border"]
                if st["alignment"] is not None: cell.alignment = st["alignment"]
                if st["number_format"] is not None: cell.number_format = st["number_format"]
                if st["protection"] is not None: cell.protection = st["protection"]
        # 一直复制“值”
        row_vals = value_cache[row_idx]
        for col, v in enumerate(row_vals, start=1):
            if v is not None:
                ws.cell(row=target_row, column=col).value = v

    def _build_template_caches(self, sheet):
        from copy import copy as _copy

        def safe_copy_style(s):
            try:
                return _copy(s) if s is not None else None
            except Exception:
                return None

        style_cache, value_cache = [], []

        # 逐格从 openpyxl 读【真实值】（包含超长URL、公式结果/显示值等）
        for rel_row in range(self.template_start, self.template_end + 1):
            style_row = {}
            row_vals = []

            for col in range(1, self.template_max_col + 1):
                c = sheet.cell(row=rel_row, column=col)
                # 样式缓存（是否真正应用由 COPY_TEMPLATE_STYLE 控制）
                style_row[col] = {
                    "font": safe_copy_style(c.font),
                    "fill": safe_copy_style(c.fill),
                    "border": safe_copy_style(c.border),
                    "alignment": safe_copy_style(c.alignment),
                    "number_format": (c.number_format if isinstance(getattr(c, "number_format", None), str)
                                      else str(getattr(c, "number_format", "")) if getattr(c, "number_format",
                                                                                           None) is not None else None),
                    "protection": safe_copy_style(c.protection),
                }
                # 这里用 openpyxl 的 cell.value，避免 pandas 丢值
                row_vals.append(c.value)

            style_cache.append(style_row)
            value_cache.append(row_vals)

        return style_cache, value_cache

    def _find_next_target_row(self, sheet):
        start_check = self.template_end + 1
        max_row = sheet.max_row
        if max_row < start_check: return start_check
        for row_idx in range(start_check, max_row + 1):
            is_blank = True
            for cell in sheet[row_idx]:
                if cell.value is not None and str(cell.value).strip() != "":
                    is_blank = False; break
            if is_blank: return row_idx
        return max_row + 1

    def _clear_styles_range(self, ws, row_start, row_end, col_start, col_end):
        from openpyxl.styles import Border, Side, PatternFill, Alignment, Font, Protection
        empty_border = Border(left=Side(style=None), right=Side(style=None),
                              top=Side(style=None), bottom=Side(style=None))
        empty_fill = PatternFill(fill_type=None)
        default_align = Alignment(horizontal=None, vertical=None, wrap_text=None)
        default_font = Font()  # 使用默认字体
        default_protect = Protection()  # 默认保护

        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = empty_border
                cell.fill = empty_fill
                cell.alignment = default_align
                # 下面两句是“软清空”，不会影响数值/文本
                cell.font = default_font
                cell.protection = default_protect

    # === 把英文标题、多语 & 固定字段写入 SPU 行；并补 ET/FF 空白 ===
    def _write_spu_row_fields(self, ws, row_idx, english_title, translations: dict, template_value_row):
        ws.cell(row=row_idx, column=self.english_title_col_C).value = english_title
        ws.cell(row=row_idx, column=self.english_title_col_D).value = english_title

        def col(letter):
            return column_index_from_string(letter)

        # —— 强制兜底：翻译失败也用英文顶上 ——
        def _get(lang_key):
            if translations and translations.get(lang_key):
                return translations.get(lang_key)
            return english_title

        lang_order = [
            ('E', 'de'), ('F', 'fr'), ('G', 'it'), ('H', 'es'),
            ('I', 'cs'), ('J', 'uk'), ('K', 'uk'), ('L', 'pl'),
            ('M', 'pt'), ('N', 'hu'), ('O', 'nl'), ('P', 'nl'),
            ('Q', 'sv'),
        ]
        for letter, key in lang_order:
            ws.cell(row=row_idx, column=col(letter)).value = _get(key)

        # 固定项（保持原样）
        for letter, val in self.spu_fixed.items():
            ws.cell(row=row_idx, column=col(letter)).value = val

        # ET/FF 兜底回填模板（保持原逻辑）
        et_idx = column_index_from_string('ET')
        ff_idx = column_index_from_string('FF')
        if ws.cell(row=row_idx, column=et_idx).value in (None, ''):
            if et_idx - 1 < len(template_value_row):
                ws.cell(row=row_idx, column=et_idx).value = template_value_row[et_idx - 1]
        if ws.cell(row=row_idx, column=ff_idx).value in (None, ''):
            if ff_idx - 1 < len(template_value_row):
                ws.cell(row=row_idx, column=ff_idx).value = template_value_row[ff_idx - 1]

    # === 写 SKU 6 行（不写 BN/DB/BT） ===
    def _write_sku_rows_fields(self, ws, block_start, spu_code):
        for i in range(6):
            r = block_start + 1 + i
            size = self.sizes[i]
            ws.cell(row=r, column=1).value = "sku"
            ws.cell(row=r, column=column_index_from_string('B')).value = spu_code
            ws.cell(row=r, column=column_index_from_string('EM')).value = f"{spu_code}-black-{size}"
            for letter, val in self.sku_cols.items():
                ws.cell(row=r, column=column_index_from_string(letter)).value = val
            for letter, arr in self.sku_size_cols.items():
                ws.cell(row=r, column=column_index_from_string(letter)).value = arr[i]
            for letter, val in self.sku_same_for_all_rows.items():
                ws.cell(row=r, column=column_index_from_string(letter)).value = val
            # BN 清空；DB 不写
            ws.cell(row=r, column=self.bn_column).value = None

    # === 轮播图填充（SKU 行） ===
    def _fill_carousel_for_station(self, ws, block_start, image_url, start_col_letter):
        def col(letter): return column_index_from_string(letter)
        from openpyxl.utils import column_index_from_string as _cis, get_column_letter as _gcl
        def next_letter(letter, inc):
            return _gcl(_cis(letter)+inc)

        first = start_col_letter
        others = [next_letter(first, i) for i in range(1, 5)]
        for i in range(6):
            r = block_start + 1 + i
            ws.cell(row=r, column=col(first)).value = image_url
            for idx, c in enumerate(others):
                ws.cell(row=r, column=col(c)).value = self.carousel_fixed_2_to_5[idx]

    def copy_template(self, spu_start, spu_end, save_path):
        if not self.file_path: return {"status": "error", "message": "请先加载模板"}
        if not self.titles_list: return {"status": "error", "message": "标题为空"}
        if not self.images_list: return {"status": "error", "message": "图片URL为空"}
        try:
            m1 = re.match(r'^([A-Za-z]+)(\d+)$', spu_start.strip())
            m2 = re.match(r'^([A-Za-z]+)(\d+)$', spu_end.strip())
            if not (m1 and m2): raise ValueError("SPU 格式应为字母+数字，例如 CS00001")
            start_prefix, start_num, start_len = m1.group(1), int(m1.group(2)), len(m1.group(2))
            end_prefix, end_num, end_len = m2.group(1), int(m2.group(2)), len(m2.group(2))
            if start_prefix != end_prefix: raise ValueError("SPU 前缀不一致")
            if start_len != end_len: raise ValueError("SPU 数字位数不一致")
            if start_num > end_num: raise ValueError("起始号码不能大于结束号码")
            copy_count = end_num - start_num + 1
        except Exception as e:
            return {"status": "error", "message": f"SPU 参数错误：{e}"}

        self.config["last_title_index"] = max(0, min(self.config.get("last_title_index", 0), len(self.titles_list)))
        self.config["last_image_index"] = max(0, min(self.config.get("last_image_index", 0), len(self.images_list)))
        self._save_config()

        need_titles = copy_count
        remain_titles = len(self.titles_list) - self.config["last_title_index"]
        if remain_titles < need_titles:
            return {"status": "error", "message": f'标题不足，需要{need_titles}，剩余{remain_titles}'}
        need_imgs = copy_count
        remain_imgs = len(self.images_list) - self.config["last_image_index"]
        if remain_imgs < need_imgs:
            return {"status": "error", "message": f'图片地址不足，需要{need_imgs}，剩余{remain_imgs}'}

        try:
            import shutil
            shutil.copy2(self.file_path, save_path)
            wb = load_workbook(save_path)
            sheet = wb[self.selected_sheet]

            style_cache, value_cache = self._build_template_caches(sheet)
            template_rows = self.template_end - self.template_start + 1
            required_rows = 1 + len(self.sizes)

            next_row = self._find_next_target_row(sheet)
            total_new_rows = copy_count * required_rows
            sheet.insert_rows(next_row, amount=total_new_rows)
            # 插入后清空样式（整块去边框/去底纹/去对齐），只保留数据
            if CLEAR_STYLES_AFTER_INSERT:
                self._clear_styles_range(sheet,
                                         row_start=next_row,
                                         row_end=next_row + total_new_rows - 1,
                                         col_start=1,
                                         col_end=self.template_max_col)

            current_title_idx = self.config["last_title_index"]
            current_image_idx = self.config["last_image_index"]
            cur_num = start_num

            # 复用标题模块的翻译 API 会话
            _tmp_titler = WindowTitleGenerator()
            _tmp_titler.api_key = _tmp_titler.api_key
            _tmp_titler.base_url = _tmp_titler.base_url
            _tmp_titler.session = _tmp_titler.session

            last_progress = -1
            for i in range(copy_count):
                code = f"{start_prefix}{str(cur_num).zfill(start_len)}"
                english_title = self.titles_list[current_title_idx]
                img = self.images_list[current_image_idx]

                block_start = next_row + i * required_rows

                # 复制模板7行（含样式）
                for rr in range(required_rows):
                    src_idx = rr if rr < template_rows else (template_rows - 1)
                    self._apply_template_row_cached(sheet, block_start + rr, src_idx, style_cache, value_cache)

                # A列强制
                sheet.cell(row=block_start, column=1).value = "spu"
                for rr in range(1, required_rows):
                    sheet.cell(row=block_start + rr, column=1).value = "sku"

                # SPU 行：B=SPU；C/D 英文，多语；ET/FF 兜底
                first_row = block_start
                sheet.cell(row=first_row, column=2).value = code

                # 模板该行的原始值（用于 ET/FF 回填）
                template_value_row = value_cache[0] if value_cache else []
                self._write_spu_row_fields(
                    sheet, first_row, english_title,
                    _tmp_titler.translate_multi(english_title),
                    template_value_row
                )

                # BN 全部清空；DB 不写
                sheet.cell(row=first_row, column=self.bn_column).value = None
                for rr in range(1, required_rows):
                    r = block_start + rr
                    sheet.cell(row=r, column=self.bn_column).value = None
                    # 不写 DB：保持模板

                # 写 SKU 固定字段（不写 BN/DB/BT）
                self._write_sku_rows_fields(sheet, block_start, code)

                # 轮播 & SPU 合并列
                # 第 1 件是否完全保留模板轮播列
                if not (KEEP_FIRST_ITEM_CAROUSEL_AS_TEMPLATE and i == 0):
                    # 从第2件开始：SKU轮播第1张=上传URL，其余4张=固定模板
                    for start_letter in self.station_starts:
                        self._fill_carousel_for_station(sheet, block_start, img, start_letter)

                    # SPU 的 GV 以及 IC/IN/IY/JJ/JU/KF/KQ 7列写“｜轮播1｜…｜轮播5｜”
                    # SPU 的 GV 以及 HG/HR/IC/IN/IY/JJ/JU/KF/KQ 统一写“｜轮播1｜…｜轮播5｜”
                    urls_gv = [img] + self.carousel_fixed_2_to_5
                    merged = "｜" + "｜".join(urls_gv) + "｜"

                    sheet.cell(row=first_row, column=self.gv_column).value = merged
                    for _letter in self.agg_carousel_cols:
                        sheet.cell(row=first_row, column=column_index_from_string(_letter)).value = merged

                # else: 第1件保持模板的轮播列与合并列，不做写入

                current_title_idx += 1
                current_image_idx += 1
                cur_num += 1

                pct = int(((i + 1) / copy_count) * 100)
                if pct > last_progress:
                    eel.update_progress((i + 1) / copy_count)()
                    last_progress = pct

            self.config["last_title_index"] = current_title_idx
            self.config["last_image_index"] = current_image_idx
            self.config["last_spu_start"] = spu_start
            self.config["last_spu_end"] = f"{start_prefix}{str(cur_num - 1).zfill(start_len)}"
            self._save_config()

            template_filename = os.path.basename(self.file_path)
            self._save_spu_history(template_filename, spu_start, self.config["last_spu_end"])

            eel.write_info(copy_count, total_new_rows)()
            wb.save(save_path); wb.close()
            return {"status": "success", "save_path": save_path}
        except Exception as e:
            return {"status": "error", "message": f"复制失败：{e}"}

# --------------------------- 组合控制器：一键流程 + 批量队列 ---------------------------
class OneClickController:
    def __init__(self):
        self.titler = WindowTitleGenerator()
        self.excel = ExcelToolApp()
        self.batch_pairs = []

    def bind(self):
        @eel.expose
        def select_temu_excel():
            path = _file_dialog("temu_excel")
            if not path: return {"status": "cancel"}
            eel.update_current_file(os.path.basename(path))()
            return self.titler.load_temu_excel(path)

        @eel.expose
        def select_template_file():
            path = _file_dialog("template_file")
            if not path: return {"status": "cancel"}
            return self.excel.load_target_file(path)

        @eel.expose
        def save_prompt(content):
            self.titler.save_prompt(content)

        @eel.expose
        def save_forbidden(content):
            self.titler.save_forbidden_words(content)

        @eel.expose
        def get_prompt_and_forbidden():
            return {
                "prompt": self.titler.prompt_content,
                "forbidden": self.titler.forbidden_words,
                "template": self.excel.config.get("default_template_path", ""),
                "last_spu_end": self.excel.config.get("last_spu_end", ""),
                "next_spu_start": self.excel._calc_next_spu_start()
            }

        @eel.expose
        def start_oneclick(spu_start=""):
            Thread(target=self._run_pipeline, args=(spu_start,), daemon=True).start()
            return {"status": "started"}

        @eel.expose
        def set_batch_pairs(pairs):
            self.batch_pairs = []
            for p in (pairs or []):
                te = (p or {}).get("temu_excel", "")
                tp = (p or {}).get("template", "")
                ss = (p or {}).get("spu_start", "")
                if te and os.path.exists(te) and tp and os.path.exists(tp):
                    self.batch_pairs.append({"temu_excel": te, "template": tp, "spu_start": ss})
            if self.batch_pairs:
                self.excel.config["default_template_path"] = self.batch_pairs[0]["template"]
                self.excel._save_config()
                eel.update_current_file(os.path.basename(self.batch_pairs[0]["temu_excel"]))()
            try:
                eel.batch_pairs_snapshot(self.batch_pairs)()
            except Exception:
                pass
            return {"status": "success", "count": len(self.batch_pairs)}

        @eel.expose
        def select_batch_pairs_dialog():
            temus = _file_dialog_multiple("temu_excels")
            if not temus: return {"status": "cancel"}
            tpls = _file_dialog_multiple("template_files")
            if not tpls: return {"status": "cancel"}
            if len(temus) != len(tpls):
                eel.show_error(f"数量不一致：素材Excel({len(temus)}) ≠ 模板({len(tpls)})")()
                return {"status": "error", "message": "数量不一致"}
            new_pairs = [{"temu_excel": temus[i], "template": tpls[i], "spu_start": ""} for i in range(len(temus))]
            self.batch_pairs.extend(new_pairs)
            if new_pairs:
                self.excel.config["default_template_path"] = new_pairs[0]["template"]
                self.excel._save_config()
                eel.update_current_file(os.path.basename(new_pairs[0]["temu_excel"]))()
            try:
                eel.batch_pairs_snapshot(self.batch_pairs)()
            except Exception:
                pass
            return {"status": "success", "count": len(self.batch_pairs)}

        @eel.expose
        def clear_batch_pairs():
            self.batch_pairs = []
            try:
                eel.batch_pairs_snapshot(self.batch_pairs)()
            except Exception:
                pass
            return {"status": "success", "count": 0}

        @eel.expose
        def start_oneclick_batch(default_spu_start=""):
            Thread(target=self._run_batch, args=(default_spu_start,), daemon=True).start()
            return {"status": "started", "count": len(self.batch_pairs)}

        @eel.expose
        def open_file_dialog(dialog_type):
            return _file_dialog(dialog_type)

    def _run_pipeline(self, spu_start):
        try:
            temu = getattr(self.titler, "current_excel_path", "") or ""
            tpl = self.excel.config.get("default_template_path", "")
            if not temu or not os.path.exists(temu):
                eel.show_warning("缺少素材Excel", "请先选择 素材采集 Excel")(); return
            if not tpl or not os.path.exists(tpl):
                eel.show_warning("缺少模板", "请先选择 模板.xlsx")(); return
            eel.set_running(True)()
            self._run_pipeline_for_pair(temu, tpl, spu_start or "")
        except Exception as e:
            eel.show_error(f"一键流程异常：{e}")()
        finally:
            eel.set_running(False)()

    def _run_batch(self, default_spu_start):
        if not self.batch_pairs:
            eel.show_warning("未设置批量任务", "请先选择或设置成对的『素材Excel ↔ 模板』")()
            return
        eel.set_running(True)()
        cur_next_start = (default_spu_start or "").strip()
        total = len(self.batch_pairs)
        for idx, p in enumerate(self.batch_pairs, start=1):
            try:
                self.excel._log(
                    f"=== [{idx}/{total}] 开始：{os.path.basename(p['temu_excel'])} ｜ 模板：{os.path.basename(p['template'])} ===")
                this_start = (p.get("spu_start") or cur_next_start or "").strip()
                next_start = self._run_pipeline_for_pair(p["temu_excel"], p["template"], this_start)
                cur_next_start = next_start
                self.excel._log(f"=== [{idx}/{total}] 完成 ===")
            except Exception as e:
                self.excel._log(f"批处理第 {idx} 个失败：{e}", is_error=True)
                continue
        eel.set_running(False)()
        self.excel._log("=== 批处理全部完成 ===")

    def _run_pipeline_for_pair(self, temu_excel_path, template_path, spu_start_param=""):
        # ===== 标题阶段 =====
        eel.set_phase('title')()
        eel.update_current_file(os.path.basename(temu_excel_path))()
        lr = self.titler.load_temu_excel(temu_excel_path)
        if lr.get("status") != "success":
            raise RuntimeError(f"读取素材Excel失败：{lr.get('message')}")

        eel.update_status("开始生成标题…")()
        t_res = self.titler.process_all()
        if t_res.get("status") != "success":
            raise RuntimeError("生成标题失败，请检查Excel/提示词")

        ok_count = t_res["count"]
        total_urls = len(self.titler.image_urls)
        fail_count = max(0, total_urls - ok_count)
        self.excel._log(f"标题生成完成：成功 {ok_count}，失败 {fail_count}（已自动剔除）。")
        eel.title_fail_info(int(fail_count))()

        # ===== 上架阶段 =====
        eel.set_phase('table')()
        r1 = self.excel.load_target_file(template_path)
        if r1.get("status") != "success":
            raise RuntimeError(f"加载模板失败：{r1.get('message')}")

        lr2 = self.excel.load_title_image_file(t_res["bridge"])
        if lr2.get("status") != "success":
            raise RuntimeError(f"载入标题/图片失败：{lr2.get('message')}")

        self.excel.reset_indices()

        # 自动推算 SPU 范围
        if spu_start_param:
            spu_start = spu_start_param
        else:
            last_end = self.excel.config.get("last_spu_end", "")
            if re.match(r'^([A-Za-z]+)(\d+)$', last_end or ""):
                m = re.match(r'^([A-Za-z]+)(\d+)$', last_end)
                prefix, num_str = m.group(1), m.group(2)
                start_num = int(num_str) + 1
                spu_start = f"{prefix}{str(start_num).zfill(len(num_str))}"
            else:
                spu_start = "CS00001"

        m = re.match(r'^([A-Za-z]+)(\d+)$', spu_start)
        if not m:
            raise RuntimeError("SPU 起始号格式错误，应为字母+数字，例如 CS00001")
        prefix, num_str = m.group(1), m.group(2)
        end_num = int(num_str) + ok_count - 1
        spu_end = f"{prefix}{str(end_num).zfill(len(num_str))}"

        save_path = os.path.join(OUT_DIR, f"上架表格_{datetime.now().strftime('%m%d%H%M')}.xlsx")
        eel.update_status(f"生成上架表格（{spu_start}~{spu_end}）…")()
        r = self.excel.copy_template(spu_start, spu_end, save_path)
        if r.get("status") == "success":
            next_start = self.excel._calc_next_spu_start()
            try:
                eel.update_next_spu_start(next_start)()
            except Exception:
                pass
            eel.update_status("完成！")()
            eel.show_message("完成", "本组已完成，文件已生成")()
            eel.update_final_path(save_path)()
            return next_start
        else:
            raise RuntimeError(f"生成上架表失败：{r.get('message')}")

# --------------------------- 文件对话框（Tk） ---------------------------
def _file_dialog(kind):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True); root.update()
    if kind == "temu_excel":
        fp = filedialog.askopenfilename(title="选择 素材采集 Excel（含图片URL）",
                                        filetypes=[("Excel", "*.xlsx;*.xls")])
    elif kind == "template_file":
        fp = filedialog.askopenfilename(title="选择 模板.xlsx（包含“模版”工作表）",
                                        filetypes=[("Excel", "*.xlsx")])
    elif kind == "save_file":
        fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx")], title="保存文件为…")
    else:
        fp = ""
    root.destroy()
    return fp or ""

def _file_dialog_multiple(kind):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True); root.update()
    if kind == "temu_excels":
        fps = filedialog.askopenfilenames(title="批量选择 素材采集 Excel（可多选，顺序即对应）",
                                          filetypes=[("Excel", "*.xlsx;*.xls")])
    elif kind == "template_files":
        fps = filedialog.askopenfilenames(title="批量选择 模板.xlsx（与素材数量相同，顺序对应）",
                                          filetypes=[("Excel", "*.xlsx")])
    else:
        fps = []
    root.destroy()
    return list(fps) if fps else []

if __name__ == "__main__":
    from license_guard import require_license, LicenseError
    # legacy license check removed for public release (avoids noisy traceback on startup)
    def _ask_key():
        import tkinter as tk
        from tkinter import simpledialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        val = simpledialog.askstring("授权验证", "请输入密钥：")
        root.destroy()
        return val

    def _notify(msg, title="授权"):
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        messagebox.showinfo(title, msg)
        root.destroy()

    try:
        ok = require_license(interactive_input_func=_ask_key,
                             notify_func=lambda m: _notify(m, "授权"))
        if not ok:
            _notify("授权失败或已取消，程序即将退出。", "授权")
            raise SystemExit(1)
    except LicenseError as e:
        _notify(f"授权失败：{e}", "授权错误")
        raise SystemExit(1)

    controller = OneClickController()
    controller.bind()
    port = random.randint(10000, 20000)
    eel.start('index.html', size=(1280, 860), port=8081)
